import os
import csv
import openpyxl
from openpyxl.styles import Protection
from openpyxl.styles import PatternFill
from text_converter import *
from settings import VERSION

HEADER_ROW = ('actor', 'japanese', 'english', 'korean', 'papago', 'comments', VERSION)


def has_header(worksheet: openpyxl.workbook.workbook.Worksheet):
    return worksheet.cell(1, 1).value == 'actor' and worksheet.cell(1, 2).value == 'japanese'


def ignore_row(first_cell):
    value = first_cell.value
    if value and (value.startswith(script_method) or
                  value.startswith('void') or
                  value == HEADER_ROW[0] or
                  value == fade_bgm_method):
        return True
    return False


class FolderConverter:
    def __init__(self, folder_path):
        self.folder_path = os.path.normpath(folder_path)
        (self.folder_directory, self.folder_name) = os.path.split(self.folder_path)

    def save_xlsx(self, sentences, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADER_ROW)
        for row_index, sentence in enumerate(sentences, 2):
            ws.append(sentence)
            if ws.cell(row_index, 1).value == play_bgm_method:
                for col in ws.iter_cols(min_row=row_index, max_row=row_index):
                    for cell in col:
                        cell.fill = PatternFill(patternType='solid', fgColor='FF9F40')
            elif ws.cell(row_index, 1).value == script_method:
                for col in ws.iter_cols(min_row=row_index, max_row=row_index):
                    for cell in col:
                        cell.fill = PatternFill(patternType='solid', fgColor='B7E3FF')
        wb.save(path)
        wb.close()

    def load_actor_translation(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        translation = {}
        for row in ws.rows:
            key = str(row[0].value)
            if key != 'None' and key in translation:
                print(f'key duplication {key}')
            translation[key] = str(row[2].value)
        wb.close()
        return translation

    def load_xlsx(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        translation = {}
        index = 0
        for row in ws.rows:
            if ignore_row(row[0]):
                continue

            if row[0].value and row[0].value == play_bgm_method:
                key = f"{index}_{row[0].value}"
                if row[3].value is not None:
                    value = row[3].value
                else:
                    value = row[1].value
            else:
                key = f"{index}_{str(row[1].value)}"
                value = str(row[3].value)

            if key != 'None' and key in translation:
                print(f'key duplication {key}')
            translation[key] = value

            index += 1
        wb.close()
        return translation

    def save_tsv(self, sentences, path):
        with open(path, 'w', encoding='utf-8', newline='') as f:
            wr = csv.writer(f, delimiter='\t')
            for sentence in sentences:
                wr.writerow(sentence)

    def load_tsv(self, path):
        with open('test.tsv', 'r', encoding='utf-8') as f:
            rdr = csv.reader(f, delimiter='\t')
            translation = {}
            for row in rdr[1:]:
                translation[row[1]] = row[3]
        return translation

    def export_text(self, format):
        converted_folder = os.path.join(self.folder_directory, self.folder_name + '_output')
        if not os.path.exists(converted_folder):
            os.mkdir(converted_folder)

        for file_name in os.listdir(self.folder_path):
            if not file_name.endswith('.txt'):
                continue

            file_path = os.path.join(self.folder_path, file_name)
            with open(file_path, 'r', encoding='utf-8') as f:
                print(f"start converting {file_name}....", end='')
                text_converter = TextConverter(f.read())
                if not text_converter.validate_text():
                    sys.exit(-1)
                sentences = text_converter.extract_text()

                file_name_only = os.path.splitext(file_name)[0]
                if format == 'xlsx':
                    self.save_xlsx(sentences, os.path.join(converted_folder, f'{file_name_only}.xlsx'))
                elif format == 'tsv':
                    self.save_tsv(sentences, os.path.join(converted_folder, f'{file_name_only}.tsv'))
                print(f"finished")

    def replace_text(self, translation_folder, actor_path, use_bgm=False):
        replaced_folder = os.path.join(self.folder_directory, self.folder_name + '_replaced')
        if not os.path.exists(replaced_folder):
            os.mkdir(replaced_folder)

        translation_base = self.load_actor_translation(actor_path)
        translation_base[None] = ''

        failed = False
        translation_folder = os.path.normpath(translation_folder)
        for file_name in os.listdir(translation_folder):
            (file_name_only, ext) = os.path.splitext(file_name)
            script_file_name = f'{file_name_only}.txt'
            script_path = os.path.join(self.folder_path, script_file_name)
            if not os.path.exists(script_path):
                continue
            print(f'start replacing {script_file_name}....', end='')
            translation = dict(translation_base)

            file_path = os.path.join(translation_folder, file_name)
            # deprecated
            if ext == '.tsv':
                full_translation = self.load_tsv(file_path)
            elif ext == '.xlsx':
                full_translation = self.load_xlsx(file_path)
            else:
                raise ModuleNotFoundError
            half_translation = dict()
            for pair in full_translation.items():
                half_translation[pair[0].translate(full_to_half_ascii).translate(custom_map)] = pair[1]
            translation.update(half_translation)

            with open(script_path, 'r', encoding='utf-8') as f:
                text_converter = TextConverter(f.read())
                replaced_text = text_converter.replace_text(translation, use_bgm=use_bgm)
                if not TextConverter(replaced_text).validate_text():
                    failed = True
                with open(os.path.join(replaced_folder, script_file_name), 'w', encoding='utf-8') as o:
                    o.write(replaced_text)

            print(f'{script_file_name} finished')
        if failed:
            print("replace finished with validation error!!!")
            sys.exit(-1)

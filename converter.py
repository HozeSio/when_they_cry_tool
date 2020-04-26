import translation_extractor
from folder_converter import *
import text_converter
import openpyxl
import urllib.request
import urllib.parse
import settings
import json
import onscript


def validate_folder(folder_path: str):
    result = True
    for file_name in os.listdir(folder_path):
        if not file_name.endswith('.txt'):
            continue

        print(f"validating {file_name}")
        file_path = os.path.join(folder_path, file_name)
        with open(file_path, 'r', encoding='utf-8') as f:
            text_converter = TextConverter(f.read())
            result &= text_converter.validate_text()

    if not result:
        print('validation error found!!!')
        sys.exit(-1)
    else:
        print('validation success.')


# deprecated
def combine_xlsx(original_folder, translated_folder):
    for file_name in os.listdir(translated_folder):
        if not file_name.endswith('.xlsx'):
            continue
        file_name = file_name.replace('kor', '')

        original_path = os.path.join(original_folder, file_name)
        if not os.path.exists(original_path):
            continue

        original_wb = openpyxl.load_workbook(original_path)
        original_ws = original_wb.active

        translated_wb = openpyxl.load_workbook(os.path.join(translated_folder, file_name))
        translated_ws = translated_wb.active

        for index, row in enumerate(translated_ws.iter_rows(), 1):
            if len(row) >= 3:
                original_ws.cell(row=index, column=4).value = row[2].value

        original_wb.save(original_path)
        original_wb.close()


# deprecated
def insert_actor_column(old_folder, actor_folder):
    for file_name in os.listdir(old_folder):
        if not file_name.endswith('.xlsx'):
            continue

        old_path = os.path.join(old_folder, file_name)
        old_wb = openpyxl.load_workbook(old_path)
        old_ws = old_wb.active

        actor_wb = openpyxl.load_workbook(os.path.join(actor_folder, file_name))
        actor_ws = actor_wb.active

        for index, row in enumerate(actor_ws.iter_rows(), 1):
            if old_ws.cell(row=index, column=2).value != row[2].value:
                print(f"{file_name} has different row at {index} {old_ws.cell(row=index, column=2).value} != {row[2].value}")
                break

        old_ws.insert_cols(2)

        for index, row in enumerate(actor_ws.iter_rows(), 1):
            old_ws.cell(row=index, column=2).value = row[1].value

        old_wb.save(old_path)
        old_wb.close()


def remove_key_column(folder_path):
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        cell = ws['A1']
        if not cell.value or not cell.value.startswith(os.path.splitext(file_name)):
            print(f'{file_name} is empty or has no key column', file=sys.stderr)
            continue

        ws.delete_cols(1)
        wb.save(file_path)
        wb.close()
        print(f'{file_name} removed')


def compare_line_count(left_folder, right_folder):
    for file_name in os.listdir(left_folder):
        left_file_path = os.path.join(left_folder, file_name)
        right_file_path = os.path.join(right_folder, file_name)

        left_wb = openpyxl.load_workbook(left_file_path)
        left_ws = left_wb.active

        right_wb = openpyxl.load_workbook(right_file_path)
        right_ws = right_wb.active

        if left_ws.max_row != right_ws.max_row:
            print(f'{file_name} has difference.')


def insert_new_rows(old_folder, new_folder):
    for file_name in os.listdir(old_folder):
        print(f'Start processing {file_name}')
        old_file_path = os.path.join(old_folder, file_name)
        new_file_path = os.path.join(new_folder, file_name)

        old_wb = openpyxl.load_workbook(old_file_path)
        old_ws = old_wb.active

        new_wb = openpyxl.load_workbook(new_file_path)
        new_ws = new_wb.active

        old_cursor = 1
        for new_cursor, row in enumerate(new_ws.iter_rows(), 1):
            same = True
            for col_index, r in enumerate(row, 1):
                same &= old_ws.cell(row=old_cursor, column=col_index).value == r.value

            first_cell = new_ws.cell(row=new_cursor, column=1)
            if same:
                pass
            elif first_cell.value and first_cell.value == text_converter.play_bgm_method:
                old_ws.insert_rows(old_cursor)
                for col_index, r in enumerate(row, 1):
                    old_ws.cell(row=old_cursor, column=col_index).value = r.value
                print(f'Added BGM line')
            #
            # elif ignore_row(first_cell) or first_cell.value and first_cell.value == text_converter.play_bgm_method:
            #     old_ws.insert_rows(old_cursor)
            #     for col_index, r in enumerate(row, 1):
            #         old_ws.cell(row=old_cursor, column=col_index).value = r.value
            #     print(f'New line added')
            # elif new_ws.cell(new_cursor, 2).value == old_ws.cell(old_cursor, 2).value:
            #     old_ws.cell(old_cursor, 1).value = new_ws.cell(new_cursor, 1).value
            #     old_ws.cell(old_cursor, 3).value = new_ws.cell(new_cursor, 3).value
            # else:
            #     print(f"Unknown cell missmatch occured at line {old_cursor}!!!")
            #     values = []
            #     for r in row:
            #         values.append(str(r.value))
            #     print(f"New row : ({','.join(values)})")
            #     values.clear()
            #     for rows in old_ws.iter_rows(old_cursor, old_cursor):
            #         for r in rows:
            #             values.append(str(r.value))
            #     print(f"Old row : ({','.join(values)})")
            #     return

            old_cursor += 1

        old_wb.save(old_file_path)
        old_wb.close()


def get_actors(folder, filter_folder):
    actors = set()
    for chapter in os.listdir(folder):
        if filter_folder and chapter != filter_folder:
            continue
        chapter_path = os.path.join(folder, chapter)
        if not os.path.isdir(chapter_path):
            continue
        for file in os.listdir(chapter_path):
            if not file.endswith('.txt'):
                continue

            file_path = os.path.join(folder, chapter, file)
            with open(file_path, 'r', encoding='utf-8') as f:
                conv = TextConverter(f.read())
                actors.update(conv.extract_actor())

    wb = openpyxl.Workbook()
    ws = wb.active
    for actor in actors:
        ws.append(actor)
    wb.save('actor_raw.xlsx')
    wb.close()


def insert_papago(folder):
    TRANSLATION_FILE = 'translation.json'
    translation_dict = {}
    if os.path.exists(TRANSLATION_FILE):
        with open(TRANSLATION_FILE, 'r', encoding='utf-8') as fd:
            translation_dict = json.load(fd)

        for file in os.listdir(folder):
            if not file.endswith('.xlsx'):
                continue

            has_update = False
            try:
                print(f"Processing {file}")
                file_path = os.path.join(folder, file)
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                if not has_header(ws):
                    ws.insert_rows(1)
                    ws.insert_cols(5)
                    for col, header in enumerate(HEADER_ROW, start=1):
                        ws.cell(1, col, header)
                    has_update = True
                for row_index, row in enumerate(ws.rows, start=1):
                    if ignore_row(row[0]):
                        continue

                    source = row[1].value
                    source = source and source.strip(' 　')
                    if not source:
                        continue

                    if source in translation_dict:
                        old_translation_cell = ws.cell(row_index, 5)
                        if old_translation_cell.value != translation_dict[source]:
                            old_translation_cell.value = translation_dict[source]
                            has_update = True
                        continue

                    encText = urllib.parse.quote(source)
                    data = f"source=ja&target=ko&text={encText}"
                    url = "https://naveropenapi.apigw.ntruss.com/nmt/v1/translation"
                    request = urllib.request.Request(url)
                    request.add_header('X-NCP-APIGW-API-KEY-ID', settings.CLIENT_ID)
                    request.add_header('X-NCP-APIGW-API-KEY', settings.CLIENT_SECRET)
                    response = urllib.request.urlopen(request, data=data.encode('utf-8'))
                    rscode = response.getcode()
                    if rscode != 200:
                        raise Exception
                    response_raw = response.read().decode('utf-8')
                    response_json = json.loads(response_raw)
                    translated_text = response_json['message']['result']['translatedText']
                    translation_dict[source] = translated_text
                    ws.cell(row_index, 5).value = translated_text
                    print(f"Translated {source} to {translated_text}")
                    has_update = True
                if has_update:
                    wb.save(file_path)
                wb.close()
            finally:
                if has_update:
                    with open(TRANSLATION_FILE, 'w', encoding='utf-8') as fd:
                        json.dump(translation_dict, fd, ensure_ascii=False)


def unique_characters(folder_path):
    characters = set()
    for chapter in os.listdir(folder_path):
        chapter_path = os.path.join(folder_path, chapter)
        if not os.path.isdir(chapter_path):
            continue
        for file in os.listdir(chapter_path):
            if not file.endswith('.xlsx'):
                continue

            wb = openpyxl.load_workbook(os.path.join(folder_path, chapter, file))
            ws = wb.active
            for index, row in enumerate(ws.rows, 1):
                korean = ws.cell(index, 4).value
                if not korean:
                    continue
                for c in korean:
                    characters.add(c)
    chars_list = list(characters)
    chars_list.sort()
    with open('characters.txt', 'w', encoding='utf-8') as fd:
        fd.write(''.join(chars_list))
    with open('old_characters.txt', 'r', encoding='utf-8') as fd:
        text = fd.read()
        old_characters = set()
        for c in text:
            old_characters.add(c)
    half_characters = set()
    for c in characters:
        half_characters.add(c.translate(text_converter.full_to_half_ascii))
    difference = half_characters.difference(old_characters)
    difference = list(difference)
    difference.sort()
    with open('difference.txt', 'w', encoding='utf-8') as fd:
        fd.write(''.join(difference))


def find_old_format(folder_path):
    for chapter in os.listdir(folder_path):
        chapter_path = os.path.join(folder_path, chapter)
        if not os.path.isdir(chapter_path):
            continue
        for file in os.listdir(chapter_path):
            if not file.endswith('.xlsx'):
                continue

            wb = openpyxl.load_workbook(os.path.join(folder_path, chapter, file))
            ws = wb.active
            if not has_header(ws):
                print(f"{os.path.join(chapter, file)} has deprecated format")


if __name__ == '__main__':
    if len(sys.argv) == 1 or sys.argv[1] == 'help':
        print(
"""\
usage: converter.py [commands]
available commands:
    export_text <Update folder>
    - export text parameter to xlsx file from the script
    replace_text <Update folder> <translation folder>
    - Replace english text to translated text
    extract_text <file_path>
    - extract text line from the onscript file and export to xlsx
    combine_xlsx <original_folder> <translated_folder>
    insert_actor_column <old_folder> <actor_folder>
"""
        )
    elif sys.argv[1] == 'export_text':
        converter = FolderConverter(sys.argv[2])
        converter.export_text('xlsx')
    elif sys.argv[1] == 'replace_text':
        converter = FolderConverter(sys.argv[2])
        converter.replace_text(sys.argv[3], sys.argv[4] if len(sys.argv) >= 5 else 'actor.xlsx')
    elif sys.argv[1] == 'validate_folder':
        validate_folder(sys.argv[2])
    elif sys.argv[1] == 'extract_text':
        extractor = translation_extractor.TextExtractor()
        extractor.extract_text(sys.argv[2])
    elif sys.argv[1] == 'combine_xlsx':
        combine_xlsx(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'insert_actor_column':
        insert_actor_column(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'remove_key_column':
        remove_key_column(sys.argv[2])
    elif sys.argv[1] == 'compare_line_count':
        compare_line_count(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'insert_new_rows':
        insert_new_rows(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'get_actors':
        get_actors(sys.argv[2], sys.argv[3] if len(sys.argv) >= 4 else None)
    elif sys.argv[1] == 'insert_papago':
        insert_papago(sys.argv[2])
    elif sys.argv[1] == 'unique_characters':
        unique_characters(sys.argv[2])
    elif sys.argv[1] == 'find_old_format':
        find_old_format(sys.argv[2])
    elif sys.argv[1] == 'export_text_onscript':
        onscript.FolderParser(sys.argv[2]).export_text(mode='onscript')
    elif sys.argv[1] == 'export_text_steam':
        onscript.FolderParser(sys.argv[2]).export_text()
    else:
        print("invalid command", file=sys.stderr)
        exit(-1)
